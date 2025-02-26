from flask import Flask, render_template, request, jsonify, send_file
import groq 
import os
import pytesseract
from pdf2image import convert_from_path
from openpyxl import Workbook
from werkzeug.utils import secure_filename
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import io
import requests
import logging

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = "uploads"

client = groq.Groq(api_key="YOUR_API_KEY")



def extract_text_from_pdf(pdf_path):
    images = convert_from_path(pdf_path)
    extracted_text = ""
    for img in images:
        extracted_text += pytesseract.image_to_string(img)
    return extracted_text
def generate_excel(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Data"
    
    for idx, (key, value) in enumerate(data.items(), start=1):
        ws.cell(row=idx, column=1, value=key)
        ws.cell(row=idx, column=2, value=value)
    
    wb.save(output_path)


@app.route('/simulate_tax_scenario', methods=['POST'])
def simulate_tax_scenario():
    data = request.json
    print("Received data:", data)  

    income = float(data.get("income", 0))
    expenses = float(data.get("expenses", 0))
    investments = float(data.get("investments", 0))
    years = int(data.get("years", 5))
    future_income = income * (1 + 0.05) ** years 
    future_expenses = expenses * (1 + 0.03) ** years  
    future_tax = max(0, (future_income - future_expenses - investments) * 0.2)

    plt.figure()
    plt.plot(range(years), [income * (1 + 0.05) ** i for i in range(years)], label="Income")
    plt.plot(range(years), [expenses * (1 + 0.03) ** i for i in range(years)], label="Expenses")
    plt.plot(range(years), [max(0, (income * (1 + 0.05) ** i - expenses * (1 + 0.03) ** i - investments) * 0.2) for i in range(years)], label="Tax Liability")
    plt.xlabel("Years")
    plt.ylabel("Amount (₹)")
    plt.title("Tax Scenario Simulation")
    plt.legend()
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    plt.close()  

    return send_file(img, mimetype='image/png')

@app.route('/monitor_tax_bracket', methods=['POST'])
def monitor_tax_bracket():
    
    mock_transactions = [
        {"amount": 5000, "date": "2023-01-01", "description": "Salary"},
        {"amount": -1000, "date": "2023-01-05", "description": "Rent"},
        {"amount": -500, "date": "2023-01-10", "description": "Groceries"},
        {"amount": 3000, "date": "2023-02-01", "description": "Freelance Income"},
        {"amount": -200, "date": "2023-02-15", "description": "Utilities"},
    ]

    income = sum(t['amount'] for t in mock_transactions if t['amount'] > 0)
    expenses = sum(abs(t['amount']) for t in mock_transactions if t['amount'] < 0)

    tax_brackets = [250000, 500000, 1000000]
    current_bracket = next((bracket for bracket in tax_brackets if income < bracket), None)
    if current_bracket:
        difference = current_bracket - income
        alert = f"You're ₹{difference} away from the {current_bracket * 0.1}% tax bracket."
    else:
        alert = "You're in the highest tax bracket."

    return jsonify({"income": income, "expenses": expenses, "alert": alert})


@app.route('/tax_saving_recommendations', methods=['POST'])
def tax_saving_recommendations():
    data = request.json
    income = float(data.get("income", 0))
    deductions = float(data.get("deductions", 0))

    query = f"Suggest tax-saving investments for an income of ₹{income} and deductions of ₹{deductions}."
    response = client.chat.completions.create(
        model="mixtral-8x7b-32768",
        messages=[{"role": "user", "content": query}]
    )
    recommendations = response.choices[0].message.content

    return jsonify({"recommendations": recommendations})





logging.basicConfig(level=logging.DEBUG)

import csv
import io

@app.route('/eco_tax_savings', methods=['POST'])
def eco_tax_savings():
    data = request.json
    app.logger.debug(f"Received data: {data}") 
    if "income" not in data:
        return jsonify({"error": "Income is required"}), 400
    income = float(data["income"])
    app.logger.debug(f"Income: {income}")
    currency_to_tax_benefit = {
        "Dong": 0.1,
        "Yuan Renminbi": 0.15,
        "Euro": 0.2,
        "Lari": 0.1,
        "Real": 0.15,
        "Dollar": 0.25,
        "Ruble": 0.1,
    }

    try:
        response = requests.get("https://my.api.mockaroo.com/eco_investments.json?key=YOUR_API_KEY")
        app.logger.debug(f"Mockaroo API response status: {response.status_code}") 
        app.logger.debug(f"Mockaroo API raw response: {response.text}")  
        response.raise_for_status() 
        csv_data = io.StringIO(response.text)
        reader = csv.DictReader(csv_data)
        eco_investments = list(reader)
        app.logger.debug(f"Mockaroo API data: {eco_investments}")  

        suggestions = []
        for investment in eco_investments:
            if "tax_benefit_claimed" not in investment:
                app.logger.warning(f"Missing tax_benefit_claimed in investment: {investment}")
                continue
            currency = investment["tax_benefit_claimed"]
            if currency not in currency_to_tax_benefit:
                app.logger.warning(f"Unknown currency: {currency}")
                continue

            tax_benefit_percentage = currency_to_tax_benefit[currency]
            tax_savings = income * tax_benefit_percentage

            suggestions.append({
                "investment_id": investment.get("investment_id", "N/A"),
                "project_location": investment.get("project_location", "N/A"),
                "tax_benefit_claimed": currency,
                "investment_status": investment.get("investment_status", "N/A"),
                "investment_type": investment.get("investment_type", "N/A"),
                "tax_savings": tax_savings
            })

        return jsonify({"suggestions": suggestions})

    except requests.exceptions.RequestException as e:
        app.logger.error(f"Mockaroo API error: {str(e)}")  
        return jsonify({"error": f"Failed to fetch eco-friendly investment data: {str(e)}"}), 500
    except Exception as e:
        app.logger.error(f"Unexpected error: {str(e)}")  
        return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500



@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    
    filename = secure_filename(file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)
    
    extracted_text = extract_text_from_pdf(file_path)
    data = {"Gross Salary": "₹10,00,000", "TDS Deducted": "₹50,000", "80C Deduction": "₹1,50,000"}
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    excel_filename = f"Form16_Data_{timestamp}.xlsx"
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
    generate_excel(data, excel_path)
    
    return jsonify({"download_link": f"/download_excel/{excel_filename}"})

@app.route('/download_excel/<filename>')
def download_excel(filename):
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename), as_attachment=True)




from datetime import datetime, timedelta

tax_deadlines = [
    {"date": "2023-03-31", "description": "Last date for filing ITR for FY 2022-23"},
    {"date": "2023-07-31", "description": "Advance tax payment deadline"},
    {"date": "2023-09-30", "description": "GSTR-9 filing deadline"},
]

def get_upcoming_deadlines():
    today = datetime.today()
    upcoming = []
    for deadline in tax_deadlines:
        deadline_date = datetime.strptime(deadline["date"], "%Y-%m-%d")
        if deadline_date >= today:
            upcoming.append(deadline)
    return upcoming
def get_regulatory_alerts():
    query = "What are the latest changes in tax laws for 2023?"
    response = client.chat.completions.create(
        model="mixtral-8x7b-32768",
        messages=[{"role": "user", "content": query}]
    )
    return response.choices[0].message.content

@app.route('/tax_alerts', methods=['GET'])
def tax_alerts():
    deadlines = get_upcoming_deadlines()
    regulatory_alerts = get_regulatory_alerts()
    return jsonify({"deadlines": deadlines, "regulatory_alerts": regulatory_alerts})



@app.route('/tax_dashboard', methods=['POST'])
def tax_dashboard():
    data = request.json
    income = float(data.get("income", 0))
    deductions = float(data.get("deductions", 0))
    investments = float(data.get("investments", 0))

    
    taxable_income = max(0, income - deductions - investments)
    tax_liability = taxable_income * 0.2  

   
    tax_savings = investments * 0.3  

    return jsonify({
        "taxable_income": taxable_income,
        "tax_liability": tax_liability,
        "tax_savings": tax_savings
    })



def get_tax_optimized_portfolio(income, risk_tolerance):
    query = f"Suggest a tax-optimized investment portfolio for an income of ₹{income} and a risk tolerance of {risk_tolerance}."
    response = client.chat.completions.create(
        model="mixtral-8x7b-32768",
        messages=[{"role": "user", "content": query}]
    )
    return response.choices[0].message.content

@app.route('/investment_portfolio', methods=['POST'])
def investment_portfolio():
    data = request.json
    income = float(data.get("income", 0))
    risk_tolerance = data.get("risk_tolerance", "medium")  

    portfolio = get_tax_optimized_portfolio(income, risk_tolerance)
    return jsonify({"portfolio": portfolio})

if __name__ == '__main__':
    app.run(debug=True)